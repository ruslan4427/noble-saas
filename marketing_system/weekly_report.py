#!/usr/bin/env python3
"""
Noble Instagram Weekly Report Generator
Collects all activity for the week and writes to:
  1. Local Excel file: output/reports/report_YYYY-WNN.xlsx
  2. Google Sheets: creates a new tab in the Noble reporting spreadsheet

Usage:
  python weekly_report.py               # current week
  python weekly_report.py --week 2026-W23  # specific week
  python weekly_report.py --setup          # first-time Google auth
"""

import argparse
import json
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

ROOT         = Path(__file__).parent
WEEK_CONTENT = Path.home() / "Desktop" / "Noble Images" / "Week Content"
OUTPUT_DIR   = ROOT / "output" / "reports"
CREDS_FILE   = ROOT / "gdrive_credentials.json"
TOKEN_FILE   = ROOT / "gdrive_token.json"
CONFIG_FILE  = ROOT / "output" / "report_config.json"

DAY_FOLDERS = [
    "Day01_Monday", "Day02_Tuesday", "Day03_Wednesday", "Day04_Thursday",
    "Day05_Friday", "Day06_Saturday", "Day07_Sunday",
]

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

TZ = ZoneInfo("America/New_York")


# ── Date helpers ──────────────────────────────────────────────────────────────

def week_range(iso_week: str | None = None):
    """Return (monday, sunday) for a given ISO week string like '2026-W23'."""
    if iso_week:
        year, week = iso_week.split("-W")
        monday = date.fromisocalendar(int(year), int(week), 1)
    else:
        today = date.today()
        monday = today - timedelta(days=today.weekday())
    return monday, monday + timedelta(days=6)


# ── Data collectors ───────────────────────────────────────────────────────────

def collect_posts(monday: date, sunday: date) -> list[dict]:
    """Read published.json from each day folder — return posts within week range."""
    posts = []
    for i, folder_name in enumerate(DAY_FOLDERS, 1):
        folder = WEEK_CONTENT / folder_name
        pub_path = folder / "published.json"
        if not pub_path.exists():
            continue
        try:
            pub = json.loads(pub_path.read_text())
        except Exception:
            continue

        caption_file = folder / "caption.txt"
        caption = caption_file.read_text(encoding="utf-8").strip()[:120] if caption_file.exists() else ""

        # Collect schedule info
        sched_time = _get_scheduled_time(i)

        for kind, media_id in pub.items():
            posts.append({
                "Day": i,
                "Folder": folder_name,
                "Type": kind.capitalize(),
                "Scheduled": sched_time,
                "Media ID": media_id,
                "Caption": caption,
                "URL": f"https://www.instagram.com/p/{media_id}/" if kind != "story" else "",
            })
    return posts


def collect_engagement(monday: date, sunday: date) -> list[dict]:
    """Read small_barbers.json and commented_posts.json."""
    rows = []
    eng_dir = ROOT / "output" / "engagement"

    # Small barbers log
    sb_path = eng_dir / "small_barbers.json"
    if sb_path.exists():
        try:
            state = json.loads(sb_path.read_text())
            for entry in state.get("results_log", []):
                d = _parse_date(entry.get("date", ""))
                if d and monday <= d <= sunday:
                    rows.append({
                        "Date": entry.get("date", ""),
                        "Account": "@" + entry.get("username", ""),
                        "Followers": entry.get("followers", ""),
                        "Hashtag": "#" + entry.get("hashtag", ""),
                        "Action": _action_label(entry),
                        "Comment Text": entry.get("comment", ""),
                        "Followed": "Yes" if entry.get("followed") else "No",
                        "Liked": "Yes" if entry.get("liked") else "No",
                    })
        except Exception:
            pass

    # Engagement run log (fallback — parse log file)
    log_path = ROOT / "output" / "engagement_run.log"
    if log_path.exists() and not rows:
        for line in log_path.read_text(encoding="utf-8").splitlines():
            if "followers)" in line and ("\"" in line or "followed" in line):
                # Parse lines like: [2026-06-04 09:15:32] @username (1234 flw) "comment" | followed
                try:
                    ts_str = line[1:20]
                    d = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S").date()
                    if monday <= d <= sunday:
                        rows.append({
                            "Date": d.isoformat(),
                            "Account": "",
                            "Followers": "",
                            "Hashtag": "",
                            "Action": line[22:].strip(),
                            "Comment Text": "",
                            "Followed": "",
                            "Liked": "",
                        })
                except Exception:
                    pass

    return rows


def collect_replies(monday: date, sunday: date) -> list[dict]:
    """Read engagement_run.log for reply entries."""
    rows = []
    log_path = ROOT / "output" / "engagement_run.log"
    if not log_path.exists():
        return rows

    for line in log_path.read_text(encoding="utf-8").splitlines():
        if "Replied to @" in line or ("@" in line and "account_type" in line.lower()):
            try:
                ts_str = line[1:20]
                d = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S").date()
                if monday <= d <= sunday:
                    rows.append({
                        "Date": d.isoformat(),
                        "Detail": line[22:].strip(),
                    })
            except Exception:
                pass
    return rows


def collect_account_stats() -> dict:
    """Fetch live account stats via instagrapi (session only, no re-login)."""
    try:
        import warnings; warnings.filterwarnings("ignore")
        from instagrapi import Client
        session_file = ROOT / "mcp_instagram" / "session.json"
        if not session_file.exists():
            return {}
        cl = Client()
        cl.load_settings(session_file)
        info = cl.user_info_by_username("noble.booking")
        return {
            "followers": info.follower_count,
            "following": info.following_count,
            "posts":     info.media_count,
        }
    except Exception as e:
        print(f"  ⚠  Account stats unavailable: {e}")
        return {}


def collect_daily_stats(monday: date, sunday: date) -> dict:
    stats_path = ROOT / "output" / "engagement" / "daily_stats.json"
    if stats_path.exists():
        try:
            return json.loads(stats_path.read_text())
        except Exception:
            pass
    return {}


def _get_scheduled_time(day: int) -> str:
    sched_dir = ROOT / "output" / "schedule"
    files = sorted(sched_dir.glob("schedule_*.json"), reverse=True)
    for f in files:
        try:
            data = json.loads(f.read_text())
            for entry in data.get("schedule", []):
                if entry.get("day") == day:
                    return entry.get("datetime_local", "")
        except Exception:
            pass
    return ""


def _parse_date(s: str) -> date | None:
    try:
        return date.fromisoformat(s[:10])
    except Exception:
        return None


def _action_label(entry: dict) -> str:
    parts = []
    if entry.get("comment"): parts.append("Commented")
    if entry.get("followed"): parts.append("Followed")
    if entry.get("liked"): parts.append("Liked")
    return " + ".join(parts) if parts else "Visited"


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(monday: date, sunday: date, posts, engagement, replies, account_stats=None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed")
        return None

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    iso_week = monday.isocalendar()
    filename = OUTPUT_DIR / f"noble_report_{monday.year}-W{iso_week.week:02d}.xlsx"

    wb = Workbook()

    GOLD   = "C9A84C"
    DARK   = "1A1208"
    CREAM  = "F5F0E8"
    WHITE  = "FFFFFF"
    LIGHT  = "FDF8F0"

    def header_style(cell, bg=DARK, fg=WHITE):
        cell.font = Font(bold=True, color=fg, size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def gold_header(cell):
        header_style(cell, bg=GOLD, fg=DARK)

    def auto_width(ws, min_w=12, max_w=50):
        for col in ws.columns:
            length = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

    def add_border(ws):
        thin = Border(
            left=Side(style="thin", color="E8DFC9"),
            right=Side(style="thin", color="E8DFC9"),
            top=Side(style="thin", color="E8DFC9"),
            bottom=Side(style="thin", color="E8DFC9"),
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Noble Instagram — Weekly Report  {monday.strftime('%b %d')} – {sunday.strftime('%b %d, %Y')}"
    ws["A1"].font = Font(bold=True, size=14, color=DARK)
    ws["A1"].fill = PatternFill("solid", fgColor=GOLD)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    stats = account_stats or {}
    metrics = [
        ("— ACCOUNT —",          ""),
        ("Followers",             stats.get("followers", "N/A")),
        ("Following",             stats.get("following", "N/A")),
        ("Total Posts",           stats.get("posts",     "N/A")),
        ("",                      ""),
        ("— ACTIVITY THIS WEEK —",""),
        ("Posts Published",       len(posts)),
        ("Accounts Engaged",      len(engagement)),
        ("Replies Given",         len(replies)),
        ("Comments Left",         sum(1 for e in engagement if "Commented" in e.get("Action", ""))),
        ("New Follows",           sum(1 for e in engagement if "Followed" in e.get("Action", ""))),
        ("Likes Left",            sum(1 for e in engagement if "Liked" in e.get("Action", ""))),
    ]

    ws.append([])
    ws.append(["Metric", "Value"])
    for cell in ws[3]:
        gold_header(cell)
    ws.row_dimensions[3].height = 22

    for label, val in metrics:
        ws.append([label, val])

    for row in ws.iter_rows(min_row=4, max_row=4 + len(metrics) - 1):
        for i, cell in enumerate(row):
            cell.fill = PatternFill("solid", fgColor=LIGHT if cell.row % 2 == 0 else WHITE)
            cell.alignment = Alignment(horizontal="left" if i == 0 else "center", vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    add_border(ws)

    # ── Sheet 2: Posts ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Posts Published")
    ws2.sheet_view.showGridLines = False

    headers = ["Day", "Folder", "Type", "Scheduled Time", "Media ID", "Caption Preview", "URL"]
    ws2.append(headers)
    for cell in ws2[1]:
        gold_header(cell)
    ws2.row_dimensions[1].height = 22

    for i, row in enumerate(posts):
        ws2.append([row.get(h.replace(" ", " "), row.get(h, "")) for h in
                    ["Day", "Folder", "Type", "Scheduled", "Media ID", "Caption", "URL"]])
        for cell in ws2[i + 2]:
            cell.fill = PatternFill("solid", fgColor=LIGHT if i % 2 == 0 else WHITE)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    auto_width(ws2)
    add_border(ws2)

    # ── Sheet 3: Engagement ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Engagement")
    ws3.sheet_view.showGridLines = False

    eng_headers = ["Date", "Account", "Followers", "Hashtag", "Action", "Comment Text", "Followed", "Liked"]
    ws3.append(eng_headers)
    for cell in ws3[1]:
        gold_header(cell)
    ws3.row_dimensions[1].height = 22

    for i, row in enumerate(engagement):
        ws3.append([row.get(h, "") for h in eng_headers])
        for cell in ws3[i + 2]:
            cell.fill = PatternFill("solid", fgColor=LIGHT if i % 2 == 0 else WHITE)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    auto_width(ws3)
    add_border(ws3)

    # ── Sheet 4: Replies ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Replies")
    ws4.sheet_view.showGridLines = False
    ws4.append(["Date", "Detail"])
    for cell in ws4[1]:
        gold_header(cell)
    ws4.row_dimensions[1].height = 22

    for i, row in enumerate(replies):
        ws4.append([row.get("Date", ""), row.get("Detail", "")])
        for cell in ws4[i + 2]:
            cell.fill = PatternFill("solid", fgColor=LIGHT if i % 2 == 0 else WHITE)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    auto_width(ws4, max_w=80)
    add_border(ws4)

    wb.save(filename)
    return filename


# ── Google Sheets upload ──────────────────────────────────────────────────────

def get_sheets_service():
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_FILE.write_text(creds.to_json())

    return build("sheets", "v4", credentials=creds)


def upload_to_sheets(monday: date, sunday: date, posts, engagement, replies, spreadsheet_id: str, account_stats=None):
    """Add a new tab to the existing Noble spreadsheet with this week's report."""
    service = get_sheets_service()
    iso_week = monday.isocalendar()
    tab_name = f"Week {monday.year}-W{iso_week.week:02d}"

    # Check if tab exists; create if not, clear if yes
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing = [s["properties"]["title"] for s in meta.get("sheets", [])]
    if tab_name not in existing:
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{"addSheet": {"properties": {"title": tab_name}}}]},
        ).execute()
    else:
        service.spreadsheets().values().clear(
            spreadsheetId=spreadsheet_id, range=f"{tab_name}!A1:Z500"
        ).execute()

    rows = []

    # Title row
    rows.append([f"Noble Instagram Report — {monday.strftime('%b %d')} to {sunday.strftime('%b %d, %Y')}"])
    rows.append([])

    # Summary
    stats = account_stats or {}
    rows.append(["SUMMARY"])
    rows.append(["— ACCOUNT —", ""])
    rows.append(["Followers",   stats.get("followers", "N/A")])
    rows.append(["Following",   stats.get("following", "N/A")])
    rows.append(["Total Posts", stats.get("posts",     "N/A")])
    rows.append([])
    rows.append(["— ACTIVITY THIS WEEK —", ""])
    rows.append(["Posts Published",  len(posts)])
    rows.append(["Accounts Engaged", len(engagement)])
    rows.append(["Comments Left",    sum(1 for e in engagement if "Commented" in e.get("Action", ""))])
    rows.append(["New Follows",      sum(1 for e in engagement if "Followed" in e.get("Action", ""))])
    rows.append(["Replies Given",    len(replies)])
    rows.append([])

    # Posts
    rows.append(["POSTS PUBLISHED"])
    rows.append(["Day", "Type", "Scheduled Time", "Media ID", "Caption Preview"])
    for p in posts:
        rows.append([p.get("Day",""), p.get("Type",""), p.get("Scheduled",""),
                     p.get("Media ID",""), p.get("Caption","")])
    rows.append([])

    # Engagement
    rows.append(["ENGAGEMENT — SMALL BARBERS TARGETED"])
    rows.append(["Date", "Account", "Followers", "Hashtag", "Action", "Comment Text"])
    for e in engagement:
        rows.append([e.get("Date",""), e.get("Account",""), e.get("Followers",""),
                     e.get("Hashtag",""), e.get("Action",""), e.get("Comment Text","")])
    rows.append([])

    # Replies
    rows.append(["REPLIES TO INCOMING COMMENTS"])
    rows.append(["Date", "Detail"])
    for r in replies:
        rows.append([r.get("Date",""), r.get("Detail","")])

    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{tab_name}!A1",
        valueInputOption="RAW",
        body={"values": rows},
    ).execute()

    print(f"  ✓ Uploaded to Google Sheets tab: {tab_name}")
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"


def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text())
        except Exception:
            pass
    return {}


def save_config(data: dict):
    CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
    CONFIG_FILE.write_text(json.dumps(data, indent=2))


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Noble weekly Instagram report")
    parser.add_argument("--week", default=None, help="ISO week e.g. 2026-W23")
    parser.add_argument("--setup", action="store_true", help="First-time Google auth + sheet config")
    parser.add_argument("--sheet-id", default=None, dest="sheet_id",
                        help="Google Spreadsheet ID (overrides saved config)")
    args = parser.parse_args()

    config = load_config()

    # ── Setup mode ────────────────────────────────────────────────────────────
    if args.setup:
        if not CREDS_FILE.exists():
            print("\n⚠️  gdrive_credentials.json not found.")
            print("Steps:")
            print("  1. Go to https://console.cloud.google.com/")
            print("  2. Create project → Enable Google Sheets API + Drive API")
            print("  3. Credentials → OAuth 2.0 Desktop app → Download JSON")
            print(f"  4. Save as: {CREDS_FILE}")
            print("  5. Run: python weekly_report.py --setup\n")
            return

        print("Opening browser for Google authorization...")
        get_sheets_service()
        print("✓ Google auth successful!")

        sheet_id = args.sheet_id or input("\nPaste your Google Spreadsheet ID (from the URL): ").strip()
        if sheet_id:
            config["spreadsheet_id"] = sheet_id
            save_config(config)
            print(f"✓ Spreadsheet ID saved: {sheet_id}")
        return

    # ── Report generation ─────────────────────────────────────────────────────
    monday, sunday = week_range(args.week)
    print(f"\nGenerating report: {monday} → {sunday}")

    posts         = collect_posts(monday, sunday)
    engagement    = collect_engagement(monday, sunday)
    replies       = collect_replies(monday, sunday)
    account_stats = collect_account_stats()

    print(f"  Posts: {len(posts)}  |  Engagement: {len(engagement)}  |  Replies: {len(replies)}")
    if account_stats:
        print(f"  Account: {account_stats['followers']} followers / {account_stats['following']} following / {account_stats['posts']} posts")

    # Build Excel
    excel_path = build_excel(monday, sunday, posts, engagement, replies, account_stats)
    if excel_path:
        print(f"  ✓ Excel saved: {excel_path}")

    # Upload to Google Sheets
    sheet_id = args.sheet_id or config.get("spreadsheet_id")
    if sheet_id and CREDS_FILE.exists():
        try:
            url = upload_to_sheets(monday, sunday, posts, engagement, replies, sheet_id, account_stats)
            print(f"  ✓ Google Sheets: {url}")
        except Exception as e:
            print(f"  ✗ Google Sheets upload failed: {e}")
            print("    Run: python weekly_report.py --setup")
    else:
        if not sheet_id:
            print("  ℹ  Google Sheets skipped — run --setup to connect")
        elif not CREDS_FILE.exists():
            print(f"  ℹ  Google Sheets skipped — {CREDS_FILE.name} not found")

    print("\nDone.")


if __name__ == "__main__":
    main()
